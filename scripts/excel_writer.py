"""
Card-Scout Excel Writer
Reads a JSON contact record from stdin, appends or updates contacts.xlsx.
Deduplicates on (full_name + company) composite key, case-insensitive.

Usage:
    echo '<json>' | python scripts/excel_writer.py
    python scripts/excel_writer.py < payload.json

Exit lines (relayed verbatim by Claude):
    RESULT: NEW CONTACT | <name> | <company> | Row N
    RESULT: UPDATED EXISTING ROW | <name> | <company> | Row N
    ERROR: <human-readable message>
"""

import sys
import json
import os
from pathlib import Path
from datetime import date

try:
    import openpyxl
    from openpyxl import load_workbook, Workbook
    from openpyxl.utils.exceptions import InvalidFileException
    from openpyxl.styles import Font, PatternFill, Alignment
except ImportError:
    print("ERROR: openpyxl is not installed. Run: pip install openpyxl")
    sys.exit(1)

XLSX_PATH = Path(__file__).parent.parent / "contacts.xlsx"

COLUMNS = [
    ("A", "Date Scanned",                        "date_scanned"),
    ("B", "Full Name",                            "full_name"),
    ("C", "Job Title",                            "job_title"),
    ("D", "Company",                              "company"),
    ("E", "Phone",                                "phone"),
    ("F", "Email",                                "email"),
    ("G", "Address",                              "address"),
    ("H", "Website",                              "website"),
    ("I", "Card Image File",                      "card_image_file"),
    ("J", "Research Status",                      "research_status"),
    ("K", "Recent Company News (with dates + source URLs)",      "recent_company_news"),
    ("L", "Recent Personal/Career News (with dates + source URLs)", "recent_personal_news"),
    ("M", "Key Company Announcements",            "key_announcements"),
    ("N", "Biggest Challenges/Concerns",          "biggest_challenges"),
    ("O", "Suggested Ice Breakers",               "ice_breakers"),
    ("P", "Sources List",                         "sources_list"),
    ("Q", "Research Caveats",                     "research_caveats"),
]

FORMULA_PREFIXES = ('=', '+', '@', '-', '\t', '\r')

REQUIRED_KEYS = {"full_name", "company"}


def safe_str(value, max_len: int = 32767) -> str:
    """Sanitize a value for safe Excel cell storage."""
    if value is None:
        return ""
    s = str(value)
    s = s.replace('\x00', '')
    if s and s[0] in FORMULA_PREFIXES:
        s = "'" + s
    if len(s) > max_len:
        s = s[:max_len - 12] + " [TRUNCATED]"
    return s


def load_or_create_workbook(path: Path):
    """Open contacts.xlsx or create it with header row. Returns (wb, ws, is_new)."""
    if path.exists():
        try:
            wb = load_workbook(path)
        except InvalidFileException:
            print(f"ERROR: {path.name} appears corrupted. Back it up and delete it to start fresh.")
            sys.exit(1)
        ws = wb.active
        return wb, ws, False
    else:
        wb = Workbook()
        ws = wb.active
        ws.title = "Contacts"
        _write_header(ws)
        return wb, ws, True


def _write_header(ws):
    """Write the header row with formatting."""
    for col_letter, header_name, _ in COLUMNS:
        col_idx = openpyxl.utils.column_index_from_string(col_letter)
        cell = ws.cell(row=1, column=col_idx, value=header_name)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="1F3864", end_color="1F3864", fill_type="solid")
        cell.alignment = Alignment(wrap_text=True, vertical="top")


def format_workbook(ws):
    """Set column widths for readability. Only called on new file creation."""
    narrow = 18
    medium = 28
    wide = 60

    widths = {
        "A": narrow,  # Date Scanned
        "B": medium,  # Full Name
        "C": medium,  # Job Title
        "D": medium,  # Company
        "E": medium,  # Phone
        "F": medium,  # Email
        "G": medium,  # Address
        "H": medium,  # Website
        "I": medium,  # Card Image File
        "J": narrow,  # Research Status
        "K": wide,    # Recent Company News
        "L": wide,    # Recent Personal News
        "M": wide,    # Key Announcements
        "N": wide,    # Biggest Challenges
        "O": wide,    # Ice Breakers
        "P": wide,    # Sources List
        "Q": medium,  # Research Caveats
    }
    for col_letter, width in widths.items():
        ws.column_dimensions[col_letter].width = width

    ws.row_dimensions[1].height = 30


def find_existing_row(ws, full_name: str, company: str):
    """Return the 1-based row index of a matching contact, or None."""
    target_name = full_name.strip().lower()
    target_company = company.strip().lower()

    for row in ws.iter_rows(min_row=2, values_only=False):
        cell_name = row[1]    # Column B, index 1
        cell_company = row[3] # Column D, index 3

        name_val = (cell_name.value or "").strip().lower()
        company_val = (cell_company.value or "").strip().lower()

        if name_val == target_name and company_val == target_company:
            return cell_name.row

    return None


def write_contact(ws, data: dict, is_new_file: bool) -> tuple[str, int]:
    """
    Write or update a contact row.
    Returns (action_string, row_number).
    action_string is 'NEW CONTACT' or 'UPDATED EXISTING ROW'.
    """
    full_name = safe_str(data.get("full_name", ""))
    company = safe_str(data.get("company", ""))

    existing_row = find_existing_row(ws, full_name, company)

    if existing_row is not None:
        row_num = existing_row
        action = "UPDATED EXISTING ROW"
    else:
        row_num = ws.max_row + 1
        if ws.max_row == 1 and ws.cell(row=1, column=1).value is None:
            row_num = 2
        action = "NEW CONTACT"

    for col_letter, _, data_key in COLUMNS:
        col_idx = openpyxl.utils.column_index_from_string(col_letter)
        value = safe_str(data.get(data_key, ""))
        cell = ws.cell(row=row_num, column=col_idx, value=value)
        cell.alignment = Alignment(wrap_text=True, vertical="top")

    return action, row_num


def validate_payload(data: dict) -> list[str]:
    """Return a list of validation errors, empty if valid."""
    errors = []
    for key in REQUIRED_KEYS:
        if not data.get(key, "").strip():
            errors.append(f"Required field '{key}' is missing or empty.")
    return errors


def main():
    raw = sys.stdin.read().strip()
    if not raw:
        print("ERROR: No JSON payload received on stdin.")
        sys.exit(1)

    try:
        data = json.loads(raw)
    except json.JSONDecodeError as e:
        print(f"ERROR: Invalid JSON — {e}")
        sys.exit(1)

    errors = validate_payload(data)
    if errors:
        for err in errors:
            print(f"ERROR: {err}")
        sys.exit(1)

    try:
        wb, ws, is_new_file = load_or_create_workbook(XLSX_PATH)
    except PermissionError:
        print(f"ERROR: Cannot open {XLSX_PATH.name} — it may be open in Excel. Close it and run again.")
        sys.exit(1)

    if is_new_file:
        format_workbook(ws)

    action, row_num = write_contact(ws, data, is_new_file)

    try:
        wb.save(XLSX_PATH)
    except PermissionError:
        print(f"ERROR: Cannot save {XLSX_PATH.name} — it may be open in Excel. Close it and run again.")
        sys.exit(1)

    full_name = safe_str(data.get("full_name", ""))
    company = safe_str(data.get("company", ""))
    print(f"RESULT: {action} | {full_name} | {company} | Row {row_num}")


if __name__ == "__main__":
    main()
